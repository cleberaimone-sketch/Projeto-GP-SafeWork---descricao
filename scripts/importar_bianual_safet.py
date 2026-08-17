#!/usr/bin/env python3
"""
Importa o relatório bianual da SafeT (xlsx) — vendas e turmas.

Uso:
    python3 scripts/importar_bianual_safet.py <caminho-do-xlsx>

Lê a aba "Vendas RD", que traz tudo que o CSV de vendas tem mais quatro coisas
que só existem aqui: Grupo, Origem do lead, Produtos e Observações.

A coluna Produtos é o que permite contar turmas. Ela é estruturada:

    SAFE T - NR35 - TRABALHADOR (P) (1x R$ 1.390,00)
    SAFE T - NR23 (P) (10x R$ 2.290,00) | SAFE T - NR07 (P) (10x R$ 2.290,00)

Cada item separado por " | " é uma turma. (P)/(E)/(S) é a modalidade e o "Nx"
é a quantidade de vagas — NÃO é multiplicador: o valor entre parênteses já é o
total da linha. Verificado batendo a soma dos itens contra o valor da venda.

Quando a venda traz "1x" trata-se de turma fechada e a lotação real não é
registrada: gravamos vagas = NULL em vez de estimar.

Requer openpyxl:  pip3 install openpyxl
Requer no ambiente: NEXT_PUBLIC_SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import sys
import urllib.request

try:
    import openpyxl
except ImportError:
    raise SystemExit('falta openpyxl — instale com: pip3 install openpyxl')

EMPRESA_SAFET = '3765298d-80f7-4bb3-8b77-2da17b064083'
ABA = 'Vendas RD'
LOTE = 100

ITEM = re.compile(
    r'SAFE\s*T\s*-\s*(?P<nome>.+?)\s*\((?P<mod>[PES])\)'
    r'(?:\s*\([^)]*\))?'
    r'\s*\((?P<qtd>\d+)x\s*R\$\s*(?P<val>[\d.,]+)\)',
    re.I,
)

# Itens que compõem o preço mas não são turma.
NAO_TREINO = re.compile(r'MARMITA|DESLOCAMENTO|HOSPEDAGEM|COFFEE|LOCA[ÇC][ÃA]O', re.I)

MODALIDADE = {'P': 'Presencial', 'E': 'EAD / Online', 'S': 'Semipresencial'}

TEMAS = [
    ('PRIMEIROS SOCORROS', 'Primeiros Socorros'), ('BRIGADA', 'Brigada de Incêndio'),
    ('CIPA', 'CIPA'), ('SIPAT', 'SIPAT'), ('ERGONOMIA', 'Ergonomia'),
    ('SALVAMENTO', 'Salvamento'), ('AÇÕES EMERGENCIAIS', 'Ações Emergenciais'),
    ('SIMULADO', 'Simulado'), ('PALESTRA', 'Palestra'),
]


def num(s: str) -> float:
    return float(str(s).replace('.', '').replace(',', '.'))


def normalizar_norma(nome: str) -> str:
    """'NR35 - TRABALHADOR' e 'NR-05 (Grau 2)' caem ambos em NR-35 / NR-05."""
    m = re.match(r'\s*NR\s*-?\s*(\d{1,2})', nome, re.I)
    if m:
        return f'NR-{int(m.group(1)):02d}'
    up = nome.upper()
    for chave, rotulo in TEMAS:
        if chave in up:
            return rotulo
    return ' '.join(nome.split())[:40]


def iso(v) -> str | None:
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    m = re.match(r'(\d{2})/(\d{2})/(\d{4})', s)
    return f'{m.group(3)}-{m.group(2)}-{m.group(1)}' if m else None


def limpar(v) -> str | None:
    if v is None:
        return None
    s = ' '.join(str(v).split())
    return None if s in ('', '—', '-', 'Nao informada', 'Não informada', 'None') else s


def extrair_nrs(texto: str) -> list[str]:
    up = (texto or '').upper()
    achadas = set()
    for m in re.finditer(r'NR\s*-?\s*(\d{1,2})', up):
        achadas.add(f'NR-{int(m.group(1)):02d}')
    for chave, rotulo in TEMAS:
        if chave in up:
            achadas.add(rotulo)
    return sorted(achadas)


def enviar(tabela: str, linhas: list[dict], url: str, chave: str, conflito: str,
           retorno: str = 'minimal') -> list[dict]:
    req = urllib.request.Request(
        f'{url}/rest/v1/{tabela}?on_conflict={conflito}',
        data=json.dumps(linhas, ensure_ascii=False).encode('utf-8'),
        method='POST',
        headers={
            'apikey': chave,
            'Authorization': f'Bearer {chave}',
            'Content-Type': 'application/json',
            'Prefer': f'resolution=merge-duplicates,return={retorno}',
        },
    )
    with urllib.request.urlopen(req, timeout=90) as resp:
        corpo = resp.read()
    return json.loads(corpo) if retorno == 'representation' and corpo else []


def buscar_ids(url: str, chave: str) -> dict[str, str]:
    """hash_dedup -> id, para amarrar as turmas às vendas."""
    mapa: dict[str, str] = {}
    passo, desloc = 1000, 0
    while True:
        req = urllib.request.Request(
            f'{url}/rest/v1/vendas_treinamento?select=id,hash_dedup'
            f'&empresa_id=eq.{EMPRESA_SAFET}&limit={passo}&offset={desloc}',
            headers={'apikey': chave, 'Authorization': f'Bearer {chave}'},
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            lote = json.loads(resp.read())
        for r in lote:
            mapa[r['hash_dedup']] = r['id']
        if len(lote) < passo:
            return mapa
        desloc += passo


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    url = (os.environ.get('NEXT_PUBLIC_SUPABASE_URL') or '').rstrip('/')
    chave = os.environ.get('SUPABASE_SERVICE_ROLE_KEY') or ''
    if not url or not chave:
        raise SystemExit('defina NEXT_PUBLIC_SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY')

    wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)
    if ABA not in wb.sheetnames:
        raise SystemExit(f'aba "{ABA}" não encontrada — abas: {wb.sheetnames}')

    linhas = list(wb[ABA].iter_rows(values_only=True))
    cab = [str(c) if c else '' for c in linhas[0]]
    regs = [dict(zip(cab, r)) for r in linhas[1:] if r and r[0]]

    vendas, turmas_por_hash = [], {}
    for r in regs:
        data = iso(r.get('Data'))
        cliente = limpar(r.get('Cliente'))
        if not data or not cliente:
            continue
        descricao = limpar(r.get('Negociação')) or ''
        valor = float(r.get('Valor') or 0)
        h = hashlib.sha1(
            f'{data}|{cliente}|{descricao}|{valor:.2f}'.encode('utf-8')
        ).hexdigest()

        produtos = limpar(r.get('Produtos')) or ''
        vendas.append({
            'empresa_id': EMPRESA_SAFET,
            'data_venda': data,
            'cliente': cliente,
            'cnpj': limpar(r.get('CNPJ')),
            'descricao': descricao or None,
            'valor': valor,
            'vendedor': limpar(r.get('Vendedor')),
            'unidade': limpar(r.get('Unidade')),
            'tipo_contrato': limpar(r.get('Tipo de contrato')),
            'forma_pagamento': limpar(r.get('Forma de pagamento')),
            'cortesia': str(r.get('Cortesia') or '').strip().lower().startswith('s'),
            'nrs': extrair_nrs(f'{descricao} {produtos}'),
            'modalidade': 'EAD / Online' if re.search(r'\bEAD\b|ONLINE', f'{descricao} {produtos}', re.I) else 'Presencial',
            'hash_dedup': h,
            'grupo': limpar(r.get('Grupo')),
            'origem_lead': limpar(r.get('Origem do lead')),
            'produtos_raw': produtos or None,
            'observacoes': limpar(r.get('Observações')),
            'responsavel_legal': limpar(r.get('Responsável legal')),
        })

        itens = []
        for i, m in enumerate(ITEM.finditer(produtos)):
            nome = m.group('nome')
            if NAO_TREINO.search(nome):
                continue
            qtd = int(m.group('qtd'))
            itens.append({
                'data_venda': data,
                'norma': normalizar_norma(nome),
                'descricao': ' '.join(nome.split())[:200],
                'modalidade': MODALIDADE.get(m.group('mod').upper(), 'Presencial'),
                # "1x" é turma fechada: a lotação não é informada, então NULL.
                'vagas': qtd if qtd > 1 else None,
                'valor': num(m.group('val')),
                'hash_dedup': hashlib.sha1(f'{h}|{i}|{nome}'.encode('utf-8')).hexdigest(),
            })
        if itens:
            turmas_por_hash[h] = itens

    print(f'{len(vendas)} vendas, {sum(len(v) for v in turmas_por_hash.values())} turmas')

    for i in range(0, len(vendas), LOTE):
        enviar('vendas_treinamento', vendas[i:i + LOTE], url, chave, 'hash_dedup')
        print(f'  vendas {min(i + LOTE, len(vendas))}/{len(vendas)}')

    ids = buscar_ids(url, chave)
    turmas = []
    orfas = 0
    for h, itens in turmas_por_hash.items():
        vid = ids.get(h)
        if not vid:
            orfas += len(itens)
            continue
        for it in itens:
            turmas.append({**it, 'venda_id': vid, 'empresa_id': EMPRESA_SAFET})

    if orfas:
        print(f'  aviso: {orfas} turmas sem venda correspondente, ignoradas')

    for i in range(0, len(turmas), LOTE):
        enviar('turmas_treinamento', turmas[i:i + LOTE], url, chave, 'hash_dedup')
        print(f'  turmas {min(i + LOTE, len(turmas))}/{len(turmas)}')

    print('concluído')


if __name__ == '__main__':
    main()
